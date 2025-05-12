import json
import random
import re
import time

import requests
from jsonpath import jsonpath
from openpyxl import workbook
from openpyxl import load_workbook

from datetime import datetime

from get_wb_place import gey_palce_pid

import uuid


def get_place(pid):
    import requests



    headers = {
        'accept': 'application/json, text/javascript, */*; q=0.01',
        'accept-language': 'zh-CN,zh;q=0.9',
        'cookie': cookie,
        'priority': 'u=1, i',
        'referer': 'https://place.weibo.com/wandermap/poinav?poiid=B2094655D464A3FF439C&lon=0&lat=0&luicode=10000011&lfid=100101B2094655D464A3FF439C',
        'sec-ch-ua': '"Chromium";v="130", "Google Chrome";v="130", "Not?A_Brand";v="99"',
        'sec-ch-ua-mobile': '?1',
        'sec-ch-ua-platform': '"Android"',
        'sec-fetch-dest': 'empty',
        'sec-fetch-mode': 'cors',
        'sec-fetch-site': 'same-origin',
        'user-agent': 'Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Mobile Safari/537.36',
        'x-requested-with': 'XMLHttpRequest',
    }

    params = {
        'poiid': pid,
    }
    print(pid)
    response = requests.get('https://place.weibo.com/wandermap/pois', params=params, headers=headers)
    data = response.json()
    print(data)
    pid = data['poiid']  # 地点ID
    name = data['name']  # 地点名称
    address = data['address']  # 地址
    lng = data['lng']  # 经度
    lat = data['lat']  # 纬度

    print('地点ID：', pid)
    print('地点名称：', name)
    print('地址：', address)
    print('经度：', lng)
    print('纬度：', lat)

    return [pid, name, address, lng, lat]


def get_data(since_id, place , ppt):
    """
    获取基础数据
    :param since_id:
    :return:
    """

    headers = {
        'accept': 'application/json, text/plain, */*',
        'accept-language': 'zh-CN,zh;q=0.9',
        'cookie': cookie,
        'mweibo-pwa': '1',
        'priority': 'u=1, i',
        'referer': 'https://m.weibo.cn/p/index?containerid=100808ad9efa2f14d42f7d14ef876725909e27_-_lbs&lcardid=frompoi&extparam=frompoi&luicode=10000011&lfid=1001018008642010000000000',
        'sec-ch-ua': '"Chromium";v="130", "Google Chrome";v="130", "Not?A_Brand";v="99"',
        'sec-ch-ua-mobile': '?1',
        'sec-ch-ua-platform': '"Android"',
        'sec-fetch-dest': 'empty',
        'sec-fetch-mode': 'cors',
        'sec-fetch-site': 'same-origin',
        'user-agent': 'Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Mobile Safari/537.36',
        'x-requested-with': 'XMLHttpRequest',
        'x-xsrf-token': '46bb4e',
    }
    # print(place)

    params = {
        'containerid': place,
        'lcardid': 'frompoi',
        'extparam': 'frompoi',
        'luicode': '10000011',
        'lfid': f'100103type=21&q={ppt}&t=',#'10000011',
        'since_id': str(since_id),
    }

    print(params)

    response = requests.get('https://m.weibo.cn/api/container/getIndex', params=params,
                            headers=headers)
    json_data = response.json()
    return json_data


def handle_text(text):
    """
    处理文本内容
    :param text:
    :return:
    """
    text = re.sub('<a.*?>|</a>|<span.*?>|</span>|<imgs.*?>|<br />|【.*?】', ' ', text)
    return text


def get_time(tt):
    # 解析时间字符串为 datetime 对象
    original_time = datetime.strptime(tt, "%a %b %d %H:%M:%S %z %Y")

    # 转换为标准时间格式：YYYY-MM-DD HH:MM:SS
    standard_time_str = original_time.strftime("%Y-%m-%d %H:%M:%S")

    return standard_time_str


def parse_data(json_data, start_date, end_date):
    """
    解析帖子发布位置
    :param json_data:
    :return:
    """
    texts = jsonpath(json_data, '$..data.cards..mblog.text')  # 文本内容
    times = jsonpath(json_data, '$..data.cards..mblog.created_at')  # 时间
    mids = jsonpath(json_data, '$..data.cards..mblog.mid')  # 博客ID
    sources = jsonpath(json_data, '$..data.cards..mblog.source')  # 来源
    reposts_count = jsonpath(json_data, '$..data.cards..mblog.reposts_count')  # 转发数
    comments_count = jsonpath(json_data, '$..data.cards..mblog.comments_count')  # 评论数
    attitudes_count = jsonpath(json_data, '$..data.cards..mblog.attitudes_count')  # 点赞数

    pid_lst = []  # 位置id
    tts_lst = []  # 时间
    ids_lst = []  # 微博ids
    sources_lst = []  # 用户名
    reposts_count_lst = []  # 转发数
    comments_count_lst = []  # 评论数
    attitudes_count_lst = []  # 点赞数
    content_lst = []  # 内容

    for i, tt, mid, source, repost, comment, attitudes in zip(texts, times, mids, sources, reposts_count,
                                                              comments_count, attitudes_count):
        try:
            containerid = re.search(r'http://weibo.com/p/100101(.*?)"', i).group(1)
            content = handle_text(i)
            if 'B20' in containerid:
                content_lst.append(content)
                pid_lst.append(containerid)
                tts_lst.append(tt)
                ids_lst.append(mid)
                sources_lst.append(source)
                attitudes_count_lst.append(attitudes)
                comments_count_lst.append(comment)
                reposts_count_lst.append(repost)
        except Exception as e:
            print('无位置')

    # 获取地方
    for content, pid, tt, mid, source, repost, comment, attitudes in zip(content_lst, pid_lst, tts_lst, ids_lst,
                                                                         sources_lst, reposts_count_lst,
                                                                         comments_count_lst, attitudes_count_lst):
        t = datetime.strptime(tt, '%a %b %d %H:%M:%S %z %Y')
        t = t.strftime('%Y-%m-%d %H:%M')
        t = datetime.strptime(t, '%Y-%m-%d %H:%M')

        new_tt = get_time(tt)

        # 如果 start_date 为空，可以设置为默认最早时间，或跳过比较
        if start_date:
            try:
                start_lst = list(map(int, start_date.split(' ')))  # 分割并转换为整数列表
                start = datetime(start_lst[0], start_lst[1], start_lst[2])
            except:
                print("起始日期格式不正确，请检查日期输入。")
                start = None
        else:
            start = datetime(1990 , 1 , 1)  # 如果没有指定开始时间，可以将 start 设置为 None 或某个默认日期

        # 如果 end_date 为空，可以设置为当前时间，或者跳过比较
        if end_date:
            try:
                end_lst = list(map(int, end_date.split(' ')))  # 分割并转换为整数列表
                end = datetime(end_lst[0], end_lst[1], end_lst[2])
            except:
                print("结束日期格式不正确，请检查日期输入。")
                end = None
        else:
            end = datetime.now()  # 如果没有指定结束时间，默认设置为当前时间

        # 判断 t 是否在指定的时间范围内
        if start <= t and t <= end:
                print("该时间在指定范围内")

                print('时间：', new_tt)
                print('内容：', content)
                print('ID：', mid)
                print('来源：', source)
                print('转发数：', repost)
                print('评论数：', comment)
                print('点赞数：', attitudes)
                place_lst = get_place(pid)  # 单独保存地理位置
                # 保存微博信息
                save_data(1, [mid, new_tt, source, repost, comment, attitudes, content] + place_lst)
                print('----------------------------------')
                time.sleep(random.randint(1, 4))
        else:
            print("不在指定范围内")


def save_data(typ, data):
    """
    保存数据
    :param typ:
    :param data:
    :return:
    """
    wb_wb.append(data)
    ws_wb.save('微博信息.xlsx')


def parse_data_wb(json_data):
    texts = jsonpath(json_data, '$..data.cards..mblog.text')
    pid_lst = []  # 位置id
    content_lst = []  # 内容
    for i in texts:
        try:
            containerid = re.search(r'http://weibo.com/p/100101(.*?)"', i).group(1)
            content = handle_text(i)

            if 'B20' in containerid:
                content_lst.append(content)
                pid_lst.append(containerid)
            print(containerid)
        except Exception as e:
            print('无位置')

    # 获取地方
    for content, pid in zip(content_lst, pid_lst):
        # print(content)
        get_place(pid)
        print('----------------------------------')


def init_sheet():
    """
    初始化表格：微博
    :return:
    """
    import uuid

    uuid = uuid.uuid4()
    try:
        ws = load_workbook(f'微博信息.xlsx')
        wb = ws.active
    except FileNotFoundError:
        ws = workbook.Workbook()
        wb = ws.active

        wb.append(
            ['微博ID', '发布时间', '来源', '转发数', '评论数', '点赞数', '微博文本', '地点ID', '地点名称', '详细地址',
             '经度', '纬度'])
        ws.save('微博信息.xlsx')

    return ws, wb


if __name__ == '__main__':
    place = input('请输入你想要获取的地方：').strip()
    num = input('请输入要爬取的页数（1页大概10条左右）：').strip()
    # choose = input('是否选择时间筛选：(1: 是 , 0: 否)')

    start = input('初始时间(格式例如：2024 11 15)：').strip()
    end = input('结束时间(格式例如：2024 11 15)：').strip()

    print(start, end)



    place_id = gey_palce_pid(place)
    print('地点id：', place_id)

    # 获取cookies
    cookies_lst = json.loads(open('cookies.json', 'r').read())
    # 初始化表格
    ws_wb, wb_wb = init_sheet()
    for i in range(2, int(num) + 2):
        cookie = random.choice(cookies_lst)
        print('cookie: ' , cookie) 
        try:  
            json_data = get_data(215, place_id , place)
            print('json_data:',json_data) 
            parse_data(json_data, start, end)
        except Exception as e:
            print(e)
            continue
        break
