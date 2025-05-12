"""
2023-5-30
"""
import ast
import json
import random
import time

import requests
from jsonpath import jsonpath
import pandas as pd
from openpyxl import workbook
import urllib3

urllib3.disable_warnings()

res = []
import json

# 提取数据
df = pd.read_excel('企业基础信息.xlsx')
lst = []
for index, rows in df.iterrows():
    dic = dict()
    dic['id'] = rows['ids']
    dic['oid'] = rows['统一社会信用代码']
    lst.append(dic)
# cookies = [
#     'arms_uid=34de0031-5b6f-4567-8a77-b3dd28bc4a41; cna=hFmlHOLVInoCAX1SeRZobwyo; xlly_s=1; account=oauth_k1%3AwnXaaW3blywVOjX0%2F7Po29q2pl7cUdUJ7D3WdtMvC40gk6g%2FK8J9LsBmy59fbsWoi6eImp7ELVr7Z4Y8%2Btd64mu3ahSLWDPpBm%2FGyPnaDfg%3D; deviceid=dad47dd1bf0342e8a5aa558cad972baf; pub_uid=1Tn4R6fToK22KVSaGmegHQ%3D%3D; isg=BImJ5jNBCUBfk_W5UMRE8PbqmLXj1n0IlYmPtCv-TXCvcqmEcyIx2HKjsNZEKhVA; l=fBIgMkbnN1suSeEKBOfwPurza77txIRAguPzaNbMi9fP935H5ihdW1aTkuLMCnGVFsOBR3l39SNMBeYBqQd-nxv9pEcBudMmnmOk-Wf..; tfstk=c4GOBJNeIHxiSLt0T5phursA66LlZPBT0trfMfUOQtTyEo1AizcowZSkurqzXeC..',
#     'arms_uid=2fd88250-4668-45cb-aed7-36a2fd83569e; dd_home_locale=zh-cn; arms_uid=39deeb24-6cf8-4e32-86f0-cf22e08be015; l=fB_QzCJeN1Rx5nBEBOfaFurza77OSIRYYuPzaNbMi9fPOI5B5_TfW1asKUL6C3GVF65yR35uBOxyBeYBq3xonxv9pH6aVuMmndLHR35..; isg=BDo6UZE0CswGpoYszHTRK5-Ai2Bc677FyrBKvEQz5k2YN9pxLHsO1QBFh8PrvDZd'
#     'arms_uid=5746a23f-dbeb-407f-add1-fae732760556; cna=03+4HPCyY2ICAbfm3/gsLYHN; xlly_s=1; tfstk=cQmCBvakJ6fQK1-lxW9Zca-JB7qPZt54FLPIdI9FqYaB5SMCi3s4GtWuX1rzeR1..; l=fBjL99keN1uZxEzEKOfaFurza77OSIRYjuPzaNbMi9fPOk5p5WgVW1asTTL9C3GVFsWpR3o-P8F6BeYBqIDvUg4tlu3ZIADmnmOk-Wf..; isg=BCQkkUwjHO7NyGiuHNKKNvs09SIWvUgnU_CfTT5FsO-y6cSzZs0Yt1pLrUFxCoB_'
#     'cna=hFmlHOLVInoCAX1SeRZobwyo; _m_h5_tk=99f9aafcae78eb8401f5476a30b0b2c9_1685624682713; _m_h5_tk_enc=e369836b33e0c9e163493731537bfef8; xlly_s=1; _samesite_flag_=true; cookie2=1ffc347a3bbe7ced4a287b020177203c; t=7fb76e2e12bf263303ab8b86750d63d6; _tb_token_=31334ee56eee9; sgcookie=E1009Kh2YpYyxGH2ypdmwST0HERhnWZsZ5rS0OO2vqgX9COs%2FVCfxkN5OqHsNgY4iEN3YKRpcYv6J7bCgC0uDSWDJXEWpU385UH6iW4kw98o0rE%3D; unb=2206787696506; uc1=pas=0&cookie21=Vq8l%2BKCLj6Hk37282g%3D%3D&cookie16=U%2BGCWk%2F74Mx5tgzv3dWpnhjPaQ%3D%3D&existShop=false&cookie14=Uoe8j2Fp1KfpPg%3D%3D&cookie15=W5iHLLyFOGW7aA%3D%3D; uc3=nk2=F5RHpC2cIoejhjw%3D&lg2=UIHiLt3xD8xYTw%3D%3D&vt3=F8dCsf50HSE7cq95Idk%3D&id2=UUphzOV5nsYb%2Bf81eg%3D%3D; csg=6188fe50; lgc=tb256721506; cancelledSubSites=empty; cookie17=UUphzOV5nsYb%2Bf81eg%3D%3D; dnk=tb256721506; skt=5905e6b25a0b51b7; existShop=MTY4NTYxNjQ0Ng%3D%3D; uc4=id4=0%40U2grF862PT9uTiwTjVJofx730Mw%2FQeJF&nk4=0%40FY4MthPoY97ivHUhM%2Bj5ABKeA4yUPQ%3D%3D; tracknick=tb256721506; _cc_=U%2BGCWk%2F7og%3D%3D; _l_g_=Ug%3D%3D; sg=663; _nk_=tb256721506; cookie1=WvSbKUmIvCaSCoZwCZArCpjV9oN31apKdqCn1Be9FSM%3D; isg=BODgX_NAkFtulyxQRB9InkcIse6y6cSzFI42n1rxrPuOVYB_AvmUQ7Zn7f1VfnyL; XSRF-TOKEN=e7835175-53dc-48cc-b0a9-0e7d5c772616',
#     'arms_uid=f65e07a8-af2d-4ab7-9685-992899c36d0a; cna=496hG0y95AQCAX1RE2enVAIV; xlly_s=1; isg=BKGhnCm3YQlsTs39sAEFBPYVsG27ThVAxR3BgQN2nagHasE8S54lEM-ozJ5sua14; l=fBElJgC7NwsWxsKQBOfaFurza77OSIRYYuPzaNbMi9fPOLCB5pUNW1a17ZT6C3GVF6PpR35uBOxyBeYBqQAonxvOEAexDmHmndLHR35..; tfstk=dYdvruYal0mD7V3tiiHu7citAduotIL2zn8QsGj0C3KJRercCcJDBlLeX1fD5m-96gT9oCvm3dKJ5H3VSmRM27dhRKg_1qCPzMALhKMsC7BysMp41qH1z7C2ZKVGmm595HftxDcntE8VQ1inxpIWuE-4vJMqtXY2u15sxDcnbfsFztcAHb8dVI_a4YpPyEs86ERRlte2lgF_1QFhHOoptBHEjkjdS-g-yRyNhaPUkWB1.'
# ]


class Spider:
    def __init__(self):
        # 保存数据
        self.ws = workbook.Workbook()
        self.wb = self.ws.active
        self.pc_USE_AGENT = [
            'Mozilla/5.0 (Windows NT 6.2; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.90 Safari/537.36',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_4) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3451.0 Safari/537.36',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.9; rv:57.0) Gecko/20100101 Firefox/57.0',
            'Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/28.0.1500.71 Safari/537.36',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.2999.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/53.0.2785.70 Safari/537.36',
            'Mozilla/5.0 (Macintosh; U; Intel Mac OS X 10.4; en-US; rv:1.9.2.2) Gecko/20100316 Firefox/3.6.2',
            'Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/44.0.2403.155 Safari/537.36 OPR/31.0.1889.174',
            'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.1; Trident/4.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; .NET CLR 1.1.4322; MS-RTC LM 8; InfoPath.2; Tablet PC 2.0)',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Safari/537.36',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/68.0.3440.106 Safari/537.36 OPR/55.0.2994.61',
            'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/535.1 (KHTML, like Gecko) Chrome/14.0.814.0 Safari/535.1',
            'Mozilla/5.0 (Macintosh; U; PPC Mac OS X; ja-jp) AppleWebKit/418.9.1 (KHTML, like Gecko) Safari/419.3',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/43.0.2357.134 Safari/537.36',
            'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.1; Trident/6.0; Touch; MASMJS)',
            'Mozilla/5.0 (X11; Linux i686) AppleWebKit/535.21 (KHTML, like Gecko) Chrome/19.0.1041.0 Safari/535.21',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3497.100 Safari/537.36',
            'Mozilla/5.0 (Windows NT 6.2; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.90 Safari/537.36',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_4) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3451.0 Safari/537.36',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.9; rv:57.0) Gecko/20100101 Firefox/57.0',
            'Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/28.0.1500.71 Safari/537.36',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.2999.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/53.0.2785.70 Safari/537.36',
            'Mozilla/5.0 (Macintosh; U; Intel Mac OS X 10.4; en-US; rv:1.9.2.2) Gecko/20100316 Firefox/3.6.2',
            'Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/44.0.2403.155 Safari/537.36 OPR/31.0.1889.174',
            'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.1; Trident/4.0; SLCC2; .NET CLR 2.0.50727; .NET CLR 3.5.30729; .NET CLR 3.0.30729; .NET CLR 1.1.4322; MS-RTC LM 8; InfoPath.2; Tablet PC 2.0)',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Safari/537.36',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/68.0.3440.106 Safari/537.36 OPR/55.0.2994.61',
            'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/535.1 (KHTML, like Gecko) Chrome/14.0.814.0 Safari/535.1',
            'Mozilla/5.0 (Macintosh; U; PPC Mac OS X; ja-jp) AppleWebKit/418.9.1 (KHTML, like Gecko) Safari/419.3',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/43.0.2357.134 Safari/537.36',
            'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.1; Trident/6.0; Touch; MASMJS)',
            'Mozilla/5.0 (X11; Linux i686) AppleWebKit/535.21 (KHTML, like Gecko) Chrome/19.0.1041.0 Safari/535.21',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3497.100 Safari/537.36',
            'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4093.3 Safari/537.36',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_5) AppleWebKit/537.36 (KHTML, like Gecko; compatible; Swurl) Chrome/77.0.3865.120 Safari/537.36',
            'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.88 Safari/537.36',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.88 Safari/537.36',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.131 Safari/537.36',
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4086.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:75.0) Gecko/20100101 Firefox/75.0',
            'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) coc_coc_browser/91.0.146 Chrome/85.0.4183.146 Safari/537.36',
            'Mozilla/5.0 (Windows; U; Windows NT 5.2; en-US) AppleWebKit/537.36 (KHTML, like Gecko) Safari/537.36 VivoBrowser/8.4.72.0 Chrome/62.0.3202.84',
            'Mozilla/5.0 (Windows NT 6.3; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.101 Safari/537.36',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.88 Safari/537.36 Edg/87.0.664.60',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.16; rv:83.0) Gecko/20100101 Firefox/83.0',
            'Mozilla/5.0 (X11; CrOS x86_64 13505.63.0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.88 Safari/537.36',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.9; rv:68.0) Gecko/20100101 Firefox/68.0',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.101 Safari/537.36',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.88 Safari/537.36',
            'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.198 Safari/537.36 OPR/72.0.3815.400',
            'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.101 Safari/537.36',
        ]
        self.h = {
            'authority': 'holmes.taobao.com',
            'accept': 'application/json, text/plain',
            'accept-language': 'zh-CN,zh;q=0.9',
            'content-type': 'application/json',
            # 'cookie': random.choice(cookies),
            'origin': 'https://www.dingtalk.com',
            'referer': 'https://www.dingtalk.com/',
            'sec-ch-ua': '"Not.A/Brand";v="8", "Chromium";v="114", "Google Chrome";v="114"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'empty',
            'sec-fetch-mode': 'cors',
            'sec-fetch-site': 'cross-site',
            'user-agent': random.choice(self.pc_USE_AGENT),
            'Connection': 'close'
        }

        # 字段一，企业信息
        self.url2 = 'https://holmes.taobao.com/web/detail/companyDetailCardBatch'
        self.p2 = {
        "dataModuleIds": [
            482,
            483,
            484,
            486,
            453,
            485
        ],
        "type": "web",
        "pageNo": 1,
        "pageSize": 10,
        "params": [
            {
                "key": "onecomp_id",
                "value": None
            }
        ]

        }
        # 首页信息
        self.url3 = 'https://holmes.taobao.com/web/corpquery/company/category3'
        self.p3 = {
        'ocid': None

        }
        # 融资历程
        self.url4 = 'http://holmes.taobao.com/web/detail/companyDetailCardBatch'
        self.p4 = {
        "dataModuleIds": [
            503,
            502,
            472,
            471
        ],
        "type": "web",
        "pageNo": 1,
        "pageSize": 10,
        "params": [
            {
                "key": "onecomp_id",
                "value": None
            }
        ]
        }

        # 严重违法
        self.url5 = 'http://holmes.taobao.com/web/detail/companyDetailCardBatch'
        self.p5 = {
        "dataModuleIds": [
            520,
            463,
            464,
            466,
            467,
            511,
            510,
            514,
            509,
            515,
            697,
            469,
            465
        ],
        "type": "web",
        "pageNo": 1,
        "pageSize": 10,
        "params": [
            {
                "key": "onecomp_id",
                "value": None
            }
        ]
        }

        # 行政处罚
        self.url6 = 'https://holmes.taobao.com/web/detail/companyDetailCardBatch'
        self.p6 = {"dataModuleIds": [460, 456, 455, 462, 457, 490, 470, 458, 495], "type": "web", "pageNo": 1,
        "pageSize": 10,
        "params": [{"key": "onecomp_id", "value": "1180716000050222313"}]}

        # 股东信息
        self.url7 = 'https://holmes.taobao.com/web/detail/companyDetailCardBatch'
        self.p7 = {"dataModuleIds": [482, 483, 484, 486, 453, 485], "type": "web", "pageNo": 1, "pageSize": 10,
        "params": [{"key": "onecomp_id", "value": "1180716059423150165"}]}

        # 判决文书
        self.url8 = 'https://holmes.taobao.com/web/detail/companyDetailCardBatch'
        self.p8 = {"dataModuleIds": [460, 456, 455, 462, 457, 490, 470, 458, 495], "type": "web", "pageNo": 1,
        "pageSize": 10,
        "params": [{"key": "onecomp_id", "value": "1180716000050222313"}]}

        # 基本数据
        self.wb.title = '企业基础信息'
        self.wb.append(['企业名称', '统一社会信用代码', '联系电话', '电子邮箱地址', '从业人数'
                           , '注册地址', '经营范围', '企业经营状态', '企业分类(上中下)'
                           , '所在区级', '所在市级', '所在省级', '注册资金'])
        # self.ws.save('企业基础信息.xlsx')
        # 融资数据
        self.wb1 = self.ws.create_sheet(title='融资轮次')  # 选择或创建新的Worksheet
        self.wb1.append(['企业统一社会信用代码', '融资轮次', '融资金额', '投资方', '发布日期'])
        # self.ws.save('融资轮次.xlsx')
        # 严重违法
        self.wb2 = self.ws.create_sheet(title='严重违法')  # 选择或创建新的Worksheet
        self.wb2.append(['企业统一社会信用代码', '列入日期', '列入严重违法名录原因', '列入决定机关', '移出日期',
                         '移出严重违法名录原因', '移出决定机关'])
        # self.ws.save('严重违法.xlsx')
        # 经营异常
        self.wb3 = self.ws.create_sheet(title='经营异常')  # 选择或创建新的Worksheet
        self.wb3.append(['企业统一社会信用代码', '列入日期', '列入经营异常名录原因', '列入决定机关', '移出日期',
                         '移出经营异常名录原因', '移出决定机关'])
        # 行政处罚
        self.wb4 = self.ws.create_sheet(title='行政处罚')  # 选择或创建新的Worksheet
        self.wb4.append(['统一社会信用代码', '决定文书号', '行政处罚种类', '决定机关', '决定日期'])

        # 股东信息
        self.wb5 = self.ws.create_sheet(title='股东信息')  # 选择或创建新的Worksheet
        self.wb5.append(['统一社会信用代码', '股东名称', '股东类型', '认缴出资额', '持股比例', '认缴出资日期'])

        # 判决文书0
        self.wb6 = self.ws.create_sheet(title='裁决文书')  # 选择或创建新的Worksheet
        self.wb6.append(['统一社会信用代码', '案件名称', '案件原由', '案件身份', '日期', '案号'])
        self.ws.save('企业上下游数据.xlsx')

        # 数据字段
        self.infomation_base = {}  # 基本数据
        self.infomation_rongzi = []  # 融资数据
        self.infomation_breaklow = []  # 严重违法
        self.infomation_abnormal = []  # 经营异常
        self.infomation_zheng = []  # 行政处罚
        self.infomation_gudong = []  # 股东信息
        self.infomation_judge = []  # 判决文书

        # 抓取基本信息


    def base_data(self, dic):
        self.p2["params"][0]['value'] = dic['id']
        self.p3['ocid'] = dic['id']
        # 判断地址
        # if '市' in place:
        #     self.infomation_base['所在市级'] = place
        #     self.infomation_base['所在省级'] = '-'
        #     self.infomation_base['所在区级'] = '-'
        # elif '区' in place:
        #     self.infomation_base['所在市级'] = '-'
        #     self.infomation_base['所在省级'] = '-'
        #     self.infomation_base['所在区级'] = place
        # elif '省' in place:
        #     self.infomation_base['所在市级'] = '-'
        #     self.infomation_base['所在省级'] = place
        #     self.infomation_base['所在区级'] = '-'
        # else:
        #     self.infomation_base['所在市级'] = '-'
        #     self.infomation_base['所在省级'] = '-'
        #     self.infomation_base['所在区级'] = '-'
        # self.infomation_base['企业分类(上中下)'] = dic['kind']

        # # # 访问基本数据1，提取数据
        # self.parse_base_data2(self.url2, self.p2)
        # time.sleep(1)
        # 访问基本数据1，提取数据
        # self.parse_base_data(self.url3, self.p3)
        # print(dic['category'], self.infomation_base['行业小类'])
        # 保存基本数据
        # self.save_data(0)
        # 爬取融资历程
        # self.p4["params"][0]['value'] = dic['id']
        # self.parse_rongzi_data(self.url4, self.p4, dic['oid'])

        # 爬取严重违法和经营异常
        self.p5["params"][0]['value'] = dic['id']
        self.parse_breaklow_data(self.url5, self.p5, dic['oid'])

        # 数据字段
        self.infomation_base.clear()
        self.infomation_rongzi.clear()
        self.infomation_breaklow.clear()
        self.infomation_abnormal.clear()
        self.infomation_judge.clear()
        self.infomation_gudong.clear()
        self.infomation_zheng.clear()


    # 爬取严重违法和经营异常
    def parse_breaklow_data(self, url, p, id):
        try:
            html_data = requests.post(url=url, headers=self.h, data=json.dumps(p),
                                      verify=False)
            time.sleep(0.5)
            if html_data.status_code == 200:
                data = html_data.json()
                print(data)
                # 严重违法
                break_lows = jsonpath(data, '$..data.8.rows[0:]')
                # 经营异常
                abnormal_run = jsonpath(data, '$..data.0.rows[0:]')
                if break_lows:
                    for j in break_lows:
                        dic = dict()
                        dic['企业统一社会信用代码'] = id
                        dic['列入日期'] = j[1]
                        dic['列入严重违法名录原因'] = j[3]
                        dic['列入决定机关'] = j[2]
                        dic['移出日期'] = '-'
                        dic['移出严重违法名录原因'] = '-'
                        dic['移出决定机关'] = '-'
                        self.infomation_breaklow.append(dic)
                else:
                    dic = dict()
                    dic['企业统一社会信用代码'] = id
                    dic['列入日期'] = '-'
                    dic['列入严重违法名录原因'] = '-'
                    dic['列入决定机关'] = '-'
                    dic['移出日期'] = '-'
                    dic['移出严重违法名录原因'] = '-'
                    dic['移出决定机关'] = '-'
                    self.infomation_breaklow.append(dic)
                if abnormal_run:
                    for i in abnormal_run:
                        dic = dict()
                        dic['企业统一社会信用代码'] = id
                        dic['列入日期'] = i[1]
                        dic['列入经营异常名录原因'] = i[2]
                        dic['列入决定机关'] = i[5]
                        dic['移出日期'] = i[3]
                        dic['移出经营异常名录原因'] = i[4]
                        dic['移出决定机关'] = i[5]
                        self.infomation_abnormal.append(dic)
                else:
                    dic = dict()
                    dic['企业统一社会信用代码'] = id
                    dic['列入日期'] = '-'
                    dic['列入经营异常名录原因'] = '-'
                    dic['列入决定机关'] = '-'
                    dic['移出日期'] = '-'
                    dic['移出经营异常名录原因'] = '-'
                    dic['移出决定机关'] = '-'
                    self.infomation_abnormal.append(dic)

                self.save_data(2)
                self.save_data(3)
        except:
            self.parse_breaklow_data(url, p, id)


    # 爬取融资数据
    def parse_rongzi_data(self, url, p, id):
        global ips
        ips = list(ips)
        try:
            ips.remove('')
        except:
            pass
        ip = random.choice(ips)
        proxy = {
            'http': f'http://{ip}',
            'https': f'https://{ip}',
        }
        # try:
        print(random.choice(ips))
        html_data = requests.post(url=url, headers=self.h, data=json.dumps(p),
                                  )
        time.sleep(0.5)
        if html_data.status_code == 200:
            data = html_data.json()
            print(data)
            rongzi_history = jsonpath(data, '$..data.0..rows[0:]')
            if rongzi_history:
                i = 0
                for detail in rongzi_history:
                    self.infomation_rongzi.append(dict())
                    self.infomation_rongzi[i]['融资轮次'] = detail[2]
                    self.infomation_rongzi[i]['融资金额'] = detail[3]
                    self.infomation_rongzi[i]['投资方'] = detail[5]
                    self.infomation_rongzi[i]['发布日期'] = detail[1]
                    self.infomation_rongzi[i]['企业统一社会信用代码'] = id
                    i += 1
            else:
                self.infomation_rongzi.append(dict())
                self.infomation_rongzi[0]['融资轮次'] = '-'
                self.infomation_rongzi[0]['融资金额'] = '-'
                self.infomation_rongzi[0]['投资方'] = '-'
                self.infomation_rongzi[0]['发布日期'] = '-'
                self.infomation_rongzi[0]['企业统一社会信用代码'] = id
            self.save_data(1)
        # except:
        #     self.parse_rongzi_data(url, p, id,proxies)
        # self.infomation_rongzi['企业统一社会信用代码'] = self.infomation['统一社会信用代码'].


    # 访问基本数据1，提取数据
    def parse_base_data(self, url, p):
        global res
        proxies = random.choice(res)
        try:
            html_data = requests.post(url=url, headers=self.h, data=json.dumps(p), proxies=proxies, verify=False)
            if html_data.status_code == 200:
                data = html_data.json()
                name = jsonpath(data, '$..data..companyName')  # 企业名称
                phone = jsonpath(data, '$..data..phoneList')  # 手机号
                email = jsonpath(data, '$..data..emailList.0')  # 邮箱
                reg_money = jsonpath(data, '$..data..registerCapital')  # 注册金额
                small_category = jsonpath(data, '$..data..industryTwo')  # 行业小雷
                # 保存数据
                self.infomation_base['企业名称'] = name[0]
                self.infomation_base['联系电话'] = phone
                self.infomation_base['电子邮箱地址'] = email
                self.infomation_base['注册资金'] = reg_money[0]
                self.infomation_base['行业小类'] = small_category[0]
        except:
            self.parse_base_data(url, p)


    # 访问基本数据2，提取数据
    def parse_base_data2(self, url, p):
        global res
        proxies = random.choice(res)
        try:
            html_data = requests.post(url=url, data=json.dumps(p), headers=self.h, proxies=proxies, verify=False)
            if html_data.status_code == 200:
                data = html_data.json()
                # print(data)
                info_key = jsonpath(data, '$..data.[0].fields..title')  # 基本信息表头字段
                detail_info = jsonpath(data, '$..data[0]..rows.0[1:19]')  # 法定代表人单独
                i = 1
                for k, v in zip(info_key[1:], detail_info):
                    if i in [5, 12, 17, 18, 4, 9]:
                        self.infomation_base[k] = v
                    i += 1
                # print(info_key)
        except:
            self.parse_base_data2(url, p)


    def save_data(self, num):
        if num == 0:
            try:
                email = ','.join(self.infomation_base['电子邮箱地址'])
            except:
                email = '-'
            try:
                phone = ','.join(self.infomation_base['联系电话'][0])
            except:
                phone = '-'
            self.wb.append([self.infomation_base['企业名称'],
                            self.infomation_base['统一社会信用代码'],
                            phone,
                            email,
                            self.infomation_base['人员规模'],
                            self.infomation_base['注册地址'],
                            self.infomation_base['经营范围'],
                            self.infomation_base['经营状态'],
                            self.infomation_base['企业分类(上中下)'],
                            self.infomation_base['所在区级'],
                            self.infomation_base['所在市级'],
                            self.infomation_base['所在省级'],
                            self.infomation_base['注册资金'],
                            ])
            self.ws.save('企业上下游数据.xlsx')
        elif num == 1:
            for i in self.infomation_rongzi:
                self.wb1.append(
                    [i['企业统一社会信用代码'], i['融资轮次'], i['融资金额'], i['投资方'], i['发布日期']])
                self.ws.save('企业上下游数据.xlsx')
        elif num == 2:
            for i in self.infomation_breaklow:
                self.wb2.append(
                    [str(i['企业统一社会信用代码']), i['列入日期'], i['列入严重违法名录原因'], i['列入决定机关']
                        , i['移出日期'], i['移出严重违法名录原因'], i['移出决定机关']])
                self.ws.save('企业上下游数据.xlsx')
        elif num == 3:
            for i in self.infomation_abnormal:
                self.wb3.append(
                    [str(i['企业统一社会信用代码']), i['列入日期'], i['列入经营异常名录原因'], i['列入决定机关']
                        , i['移出日期'], i['移出经营异常名录原因'], i['移出决定机关']])
                self.ws.save('企业上下游数据.xlsx')
        elif num == 4:
            for i in self.infomation_zheng:
                self.wb4.append(
                    [str(i['企业统一社会信用代码']), i['决定文书号'], i['行政处罚种类'], i['决定机关']
                        , i['决定日期']])
                self.ws.save('企业上下游数据.xlsx')
        elif num == 5:
            for i in self.infomation_gudong:
                self.wb5.append(
                    [str(i['企业统一社会信用代码']), i['股东名称'], i['股东类型'], i['认缴出资额']
                        , i['持股比例'], i['认缴出资日期']])
                self.ws.save('企业上下游数据.xlsx')
        elif num == 6:
            for i in self.infomation_judge:
                self.wb6.append(
                    [str(i['企业统一社会信用代码']), i['案件名称'], i['案件原由'], i['案件身份']
                        , i['日期'], i['案号']])
                self.ws.save('企业上下游数据.xlsx')


    def main(self):
        global ips
        # 打开表格，获取里面的信息
        sign = 0
        for k in lst:
            print('当前id: ', k['id'])
            k['id'] = int(k['id'])
            if sign == 10:
                ips = requests.get(
                    url='http://api.66daili.cn/API/GetSecretProxy/?orderid=l85l2OO4923Ol972824&num=20&token=66daili&format=text&line_separator=win&protocol=http&region=domestic').text
                ips = ips.split('\r\n')
                sign = 0
            sign += 1
            self.base_data(k)


if __name__ == '__main__':
    ips = requests.get(
        url='http://api.66daili.cn/API/GetSecretProxy/?orderid=l85l2OO4923Ol972824&num=20&token=66daili&format=text&line_separator=win&protocol=http&region=domestic').text
    ips = ips.split('\r\n')
    spider = Spider()
    spider.main()
