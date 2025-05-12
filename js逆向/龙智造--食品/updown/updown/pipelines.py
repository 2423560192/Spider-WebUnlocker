# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html


# useful for handling different item types with a single interface
from itemadapter import ItemAdapter
from openpyxl import workbook
import json
from updown.items import Rongzi,Baseinfo,Crime,Manage,Administration,Judge,Shareholder
import pymysql
# 保存企业基础信息
class UpdownPipeline:
    def __init__(self):
        # 链接数据库
        self.db = pymysql.connect(
            user='root',
            password='root',
            database='panco',
            host='127.0.0.1',
            port=3306,
        )
        self.cursor = self.db.cursor()
        # 创建数据库表
        try:
            sql = 'create table 企业基础信息(id int not null auto_increment primary key ,市级 varchar(255),省级 varchar(255),区级 varchar(255),状态 varchar(255),分类 varchar(255),ids varchar(255),电话 varchar(255),邮箱 varchar(255),企业名称 varchar(255),注册金额 varchar(255),企业简介 varchar(255),行业大类 varchar(255),行业小类 varchar(255),地址 varchar(255),统一社会信用代码 varchar(255),从业人数 varchar(255),经营范围 varchar(255))'
            self.cursor.execute(sql)
        except Exception as e:
            print('表已存在!', e)

        # self.ws = workbook.Workbook()
        # self.wb = self.ws.active
        # self.wb.append(['市级', '省级','区级','状态','分类','ids', '电话','邮箱','企业名称','注册金额',
        #                 '企业简介','行业大类','行业小类','地址', '统一社会信用代码','从业人数','经营范围'])
        # self.ws.save('零食企业数据.xlsx')

    def process_item(self, item, spider):
        if isinstance(item, Baseinfo):
            item = dict(item)
            print(item)
            info = [item['city'],item['prevence'],item['area'],item['state'],item['category'],item['now_id']
                ,item['phone'],item['email'],item['name'],item['reg_money'],item['entBrief'],
                    item['big_category'],item['small_category'],item['addr'],
                    item['oid'],item['num'],item['business']]
            sql = 'insert into 企业基础信息(市级, 省级,区级,状态,分类,ids, 电话,邮箱,企业名称,注册金额,企业简介,行业大类,行业小类,地址, 统一社会信用代码,从业人数,经营范围) values ("%s","%s","%s","%s","%s","%s","%s","%s","%s","%s","%s","%s","%s","%s","%s","%s","%s")'
            self.cursor.execute(sql, info)
            self.db.commit()
            # self.wb.append(info)
            # self.ws.save('零食企业数据.xlsx')

        return item

# 保存融资信息
class RongziInfo:
    def __init__(self):
        # # 保存数据
        # self.ws = workbook.Workbook()
        # self.wb = self.ws.active
        # self.wb.append(['企业统一社会信用代码','融资轮次','融资金额','投资方','发布日期'])
        # self.ws.save('零食融资数据.xlsx')

        # 链接数据库
        self.db = pymysql.connect(
            user='root',
            password='root',
            database='panco',
            host='127.0.0.1',
            port=3306,
        )
        self.cursor = self.db.cursor()
        # 创建数据库表
        try:
            sql = 'create table 企业融资信息(id int not null auto_increment primary key ,企业统一社会信用代码 varchar(255),融资轮次 varchar(255),融资金额 varchar(255),投资方 varchar(255),发布日期 varchar(255))'
            self.cursor.execute(sql)
        except Exception as e:
            print('表已存在!', e)

    def process_item(self, item, spider):
        if isinstance(item,Rongzi):
            item = dict(item)
            print(item)
            info = [item['oid'],item['rongzi_round'],item['rongzi_money'],item['investor'],item['date']]
            # self.wb.append(info)
            # self.ws.save('零食融资数据.xlsx')
            sql = 'insert into 企业融资信息(企业统一社会信用代码,融资轮次,融资金额,投资方,发布日期) values ("%s","%s","%s","%s","%s")'
            self.cursor.execute(sql, info)
            self.db.commit()
        return item

# 保存严重违法
class CrimeInfo:
    def __init__(self):
        # # 保存数据
        # self.ws = workbook.Workbook()
        # self.wb = self.ws.active
        # self.wb.append(['企业统一社会信用代码','列入日期','列入严重违法名录原因','列入决定机关','移出日期','移出严重违法名录原因','移出决定机关'])
        # self.ws.save('零食违法数据.xlsx')

        # 链接数据库
        self.db = pymysql.connect(
            user='root',
            password='root',
            database='panco',
            host='127.0.0.1',
            port=3306,
        )
        self.cursor = self.db.cursor()
        # 创建数据库表
        try:
            sql = 'create table 企业违法信息(id int not null auto_increment primary key,企业统一社会信用代码 varchar(255) ,列入日期 varchar(255),列入严重违法名录原因 varchar(255),列入决定机关 varchar(255),移出日期 varchar(255),移出严重违法名录原因 varchar(255),移出决定机关 varchar(255))'
            self.cursor.execute(sql)
        except Exception as e:
            print('表已存在!', e)



    def process_item(self, item, spider):
        if isinstance(item,Crime):
            item = dict(item)
            print(item)
            info = [item['oid'],item['date_to'],item['reason_to'],item['decison_to'],item['dete_out'],item['reason_out'],item['decison_out']]
            # self.wb.append(info)
            # self.ws.save('零食违法数据.xlsx')
            sql = 'insert into 企业违法信息(企业统一社会信用代码,列入日期,列入严重违法名录原因,列入决定机关,移出日期,移出严重违法名录原因,移出决定机关) values ("%s","%s","%s","%s","%s","%s","%s")'
            self.cursor.execute(sql, info)
            self.db.commit()

        return item

# 保存经营异常
class ManageInfo:
    def __init__(self):
        # # 保存数据
        # self.ws = workbook.Workbook()
        # self.wb = self.ws.active
        # self.wb.append(['企业统一社会信用代码','列入日期','列入严重违法名录原因','列入决定机关','移出日期','移出严重违法名录原因','移出决定机关'])
        # self.ws.save('零食经营数据.xlsx')

        # 链接数据库
        self.db = pymysql.connect(
            user='root',
            password='root',
            database='panco',
            host='127.0.0.1',
            port=3306,
        )
        self.cursor = self.db.cursor()
        # 创建数据库表
        try:
            sql = 'create table 企业经营信息(id int not null auto_increment primary key,企业统一社会信用代码 varchar(255) ,列入日期 varchar(255),列入严重违法名录原因 varchar(255),列入决定机关 varchar(255),移出日期 varchar(255),移出严重违法名录原因 varchar(255),移出决定机关 varchar(255))'
            self.cursor.execute(sql)
        except Exception as e:
            print('表已存在!', e)

    def process_item(self, item, spider):
        if isinstance(item,Manage):
            item = dict(item)
            print(item)
            info = [item['oid'],item['date_to'],item['reason_to'],item['decison_to'],item['dete_out'],item['reason_out'],item['decison_out']]
            # self.wb.append(info)
            # self.ws.save('零食经营数据.xlsx')
            sql = 'insert into 企业经营信息(企业统一社会信用代码,列入日期,列入严重违法名录原因,列入决定机关,移出日期,移出严重违法名录原因,移出决定机关) values ("%s","%s","%s","%s","%s","%s","%s")'
            self.cursor.execute(sql, info)
            self.db.commit()
        return item

# 保存行政处罚
class AdministrationInfo:
    def __init__(self):
        # 保存数据
        # self.ws = workbook.Workbook()
        # self.wb = self.ws.active
        # self.wb.append(['企业统一社会信用代码','决定文书号','行政处罚种类','决定机关','决定日期'])
        # self.ws.save('零食行政数据.xlsx')

        # 链接数据库
        self.db = pymysql.connect(
            user='root',
            password='root',
            database='panco',
            host='127.0.0.1',
            port=3306,
        )
        self.cursor = self.db.cursor()
        # 创建数据库表
        try:
            sql = 'create table 企业行政信息(id int not null auto_increment primary key,企业统一社会信用代码 varchar(255) ,决定文书号 varchar(255),行政处罚种类 varchar(255),决定机关 varchar(255),决定日期 varchar(255))'
            self.cursor.execute(sql)
        except Exception as e:
            print('表已存在!', e)

    def process_item(self, item, spider):
        if isinstance(item,Administration):
            item = dict(item)
            print(item)
            info = [item['oid'],item['num'],item['kind'],item['decison_to'],item['date']]
            # self.wb.append(info)
            # self.ws.save('零食行政数据.xlsx')
            sql = 'insert into 企业行政信息(企业统一社会信用代码,决定文书号,行政处罚种类,决定机关,决定日期) values ("%s","%s","%s","%s","%s")'
            self.cursor.execute(sql, info)
            self.db.commit()
        return item

# 保存裁决文书
class JudgeInfo:
    def __init__(self):
        # # 保存数据
        # self.ws = workbook.Workbook()
        # self.wb = self.ws.active
        # self.wb.append(['企业统一社会信用代码','案件名称','案件原由','案件身份','日期','案号'])
        # self.ws.save('零食裁决数据.xlsx')

        # 链接数据库
        self.db = pymysql.connect(
            user='root',
            password='root',
            database='panco2',
            host='127.0.0.1',
            port=3306,
        )
        self.cursor = self.db.cursor()
        # 创建数据库表
        try:
            sql = 'create table 企业裁决信息(id int not null auto_increment primary key,企业统一社会信用代码 varchar(255) ,案件名称 varchar(255),案件原由 varchar(255),案件身份 varchar(255),日期 varchar(255),案号 varchar(255))'
            self.cursor.execute(sql)
        except Exception as e:
            print('表已存在!', e)

    def process_item(self, item, spider):
        if isinstance(item,Judge):
            item = dict(item)
            print(item)
            info = [item['oid'],item['name'],item['reason'],item['id'],item['date'],item['num']]
            # self.wb.append(info)
            # self.ws.save('零食裁决数据.xlsx')
            sql = 'insert into 企业裁决信息(企业统一社会信用代码,案件名称,案件原由,案件身份,日期,案号) values ("%s","%s","%s","%s","%s","%s")'
            self.cursor.execute(sql, info)
            self.db.commit()
        return item

# 保存股东
class ShareholderInfo:
    def __init__(self):
        # 保存数据
        # self.ws = workbook.Workbook()
        # self.wb = self.ws.active
        # self.wb.append(['企业统一社会信用代码','股东','持股比例','认缴出资额','发布日期'])
        # self.ws.save('零食股东数据.xlsx')

        # 链接数据库
        self.db = pymysql.connect(
            user='root',
            password='root',
            database='panco',
            host='127.0.0.1',
            port=3306,
        )
        self.cursor = self.db.cursor()
        # 创建数据库表
        try:
            sql = 'create table 企业股东信息(id int not null auto_increment primary key,企业统一社会信用代码 varchar(255) ,股东 varchar(255),持股比例 varchar(255),认缴出资额 varchar(255),发布日期 varchar(255))'
            self.cursor.execute(sql)
        except Exception as e:
            print('表已存在!', e)

    def process_item(self, item, spider):
        if isinstance(item,Shareholder):
            item = dict(item)
            print(item)
            info = [item['oid'],item['name'],item['ratio'],item['money'],item['date']]
            sql = 'insert into 企业股东信息(企业统一社会信用代码,股东,持股比例,认缴出资额,发布日期) values ("%s","%s","%s","%s","%s")'
            self.cursor.execute(sql, info)
            self.db.commit()
            # self.wb.append(info)
            # self.ws.save('零食股东数据.xlsx')
        return item